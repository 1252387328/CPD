import pandas as pd
import xlrd
import ast
from openpyxl import Workbook

# Load the Excel file
file_name = 'res-16-300-100'
file_path = f'res/{file_name}.xlsx'

readbook = xlrd.open_workbook(file_path)
sheet = readbook.sheet_by_index(1)
nrowsmax = sheet.nrows

hv_list = []
pd_list = []
igd_list = []

# 假设每组三行：HV，PD，IGD
for i in range(0, int(nrowsmax / 3)):
    hv_list.append(ast.literal_eval(sheet.cell(3 * i, 4).value))
    pd_list.append(ast.literal_eval(sheet.cell(3 * i + 1, 4).value))
    igd_list.append(ast.literal_eval(sheet.cell(3 * i + 2, 4).value))

# 标题生成
def generate_headers(metric):
    headers = []
    for i in range(1, 5):  # CPD-1 到 CPD-4
        for var in ['A', 'B', 'C', 'D']:
            headers.append(f'CPD-{i}-{var}-{metric}')
    return headers

hv_headers = generate_headers('HV')
pd_headers = generate_headers('PD')
igd_headers = generate_headers('IGD')

# 写入工作簿
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

def write_sheet(ws, data_list, headers):
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col+1, value=header)
    for row in range(len(data_list[0])):
        for group in range(4):  # CPD-1 to CPD-4
            for var in range(4):  # A to D
                col = group * 4 + var
                ws.cell(row=row+2, column=col+1, value=data_list[group][row][var])

# HV sheet
ws_hv = wb.create_sheet('HV')
write_sheet(ws_hv, hv_list, hv_headers)

# PD sheet
ws_pd = wb.create_sheet('PD')
write_sheet(ws_pd, pd_list, pd_headers)

# IGD sheet
ws_igd = wb.create_sheet('IGD')
write_sheet(ws_igd, igd_list, igd_headers)

# 保存文件
wb.save(f"res/{file_name}-p.xlsx")